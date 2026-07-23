from openpyxl import load_workbook
import json
import os

# Paths are relative to this script, so it runs from anywhere
# (VS Code, terminal, etc.). Drop the source spreadsheets in ./data/
# and the generated catalogue is written to ./site/prices.json.
HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, 'data')
SITE_DIR = os.path.join(HERE, 'site')

SW = os.path.join(DATA_DIR, 'Master_price_list_Stonewater_com_pro_wef_01_04_2026.xlsx')
KA = os.path.join(DATA_DIR, 'Kasper_Professional_Audio_Price_List_2026.xlsx')

def num(v):
    if isinstance(v, bool): return None
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        s = v.replace(',', '').replace('₹', '').strip()
        try: return float(s)
        except: return None
    return None

def clean(v):
    if isinstance(v, float) and v.is_integer(): return int(v)
    return v

def assert_headers(ws, header_row, expect):
    """Fail loudly if a mapped column index no longer holds the expected header.

    Inserting a column into a source sheet shifts every later index by one.
    Without this check the build still 'succeeds' and silently writes wrong
    values into prices.json -- e.g. MRP reading the List+Tax column. Each
    expected value is matched case-insensitively as a prefix, so trailing
    whitespace or a renamed suffix in the sheet will not trip it.
    """
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[header_row]
    problems = []
    for idx, want in sorted(expect.items()):
        got = hdr[idx] if idx < len(hdr) else None
        got_s = '' if got is None else str(got).strip()
        if not got_s.lower().startswith(want.strip().lower()):
            problems.append(
                f"  col {idx}: expected header starting {want!r}, found {got_s!r}")
    if problems:
        raise SystemExit(
            f"\nCOLUMN MAP MISMATCH in sheet {ws.title!r}:\n"
            + "\n".join(problems)
            + "\n\nA column was probably inserted, removed or renamed. Update the"
              "\nfield indices and the expect={...} map in build_prices.py to match."
              "\nNothing was written -- prices.json is unchanged.\n")


def build_list(ws, *, header_row, model, desc, hsn, fields, labels, roles,
               percent_keys, top_levels, expect=None):
    if expect:
        assert_headers(ws, header_row, expect)
    rows = list(ws.iter_rows(values_only=True))
    price_cols = [c for c, _ in fields.values()]
    cats = []
    def ensure_top(name):
        cats.append({'name': name, 'subcategories': []})
    def ensure_sub(name):
        if not cats: ensure_top('')
        cats[-1]['subcategories'].append({'name': name, 'products': []})
    for r in rows[header_row + 1:]:
        nonempty = [c for c in r if c not in (None, '')]
        if not nonempty: continue
        d = r[desc] if (desc is not None and len(r) > desc) else None
        has_price = any(((num(r[i]) or 0) != 0) for i in price_cols if i < len(r))
        if (d in (None, '')) and not has_price:
            label = str(next(c for c in r if c not in (None, ''))).strip()
            if top_levels and label.lower() in top_levels:
                ensure_top(label)
            else:
                ensure_sub(label)
            continue
        mv = r[model] if len(r) > model else None
        if mv in (None, ''): continue
        if not cats: ensure_top('')
        if not cats[-1]['subcategories']: ensure_sub('')
        prices = {}
        for key, (col, kind) in fields.items():
            v = num(r[col]) if col < len(r) else None
            if v is not None and v != 0:
                prices[key] = round(v, 2)
        cats[-1]['subcategories'][-1]['products'].append({
            'model': str(mv).strip(),
            'description': '' if d in (None, '') else str(d).strip(),
            'hsn': (clean(r[hsn]) if (hsn is not None and len(r) > hsn and r[hsn] not in (None, '')) else None),
            'prices': prices,
        })
    for c in cats:
        c['subcategories'] = [s for s in c['subcategories'] if s['products']]
    cats = [c for c in cats if c['subcategories']]
    return {'labels': labels, 'roles': roles, 'percentKeys': percent_keys, 'categories': cats}

# ---------------- Stonewater ----------------
swwb = load_workbook(SW, data_only=True)
SW_TOP = {'commercial', 'pro audio'}

praveen = build_list(swwb['Praveen Price List'], header_row=2, model=1, desc=2, hsn=3,
    fields={'listPrice':(4,'inr'),'dealer':(5,'inr'),'subdealer':(6,'inr'),'distInclTax':(7,'inr'),
            'msrp':(8,'inr'),'msrp35':(9,'inr'),'msrp30':(10,'inr'),'msrp20':(11,'inr'),
            'msrp15':(12,'inr'),'dealerMargin':(13,'pct'),'msrpMargin':(14,'pct')},
    labels={'listPrice':'List Price +18%','dealer':'Dealer','subdealer':'Sub-dealer',
            'distInclTax':'Distributor (tax incl.)','msrp':'MSRP','msrp35':'MSRP −35%',
            'msrp30':'MSRP −30%','msrp20':'MSRP −20%','msrp15':'MSRP −15%',
            'dealerMargin':'Dealer margin','msrpMargin':'MSRP margin'},
    roles={'customer':['msrp'],'dealer':['dealer','msrp'],'subdealer':['subdealer','msrp'],
           'internal':['msrp','msrp35','msrp30','msrp20','msrp15','dealer','subdealer','listPrice','distInclTax','dealerMargin','msrpMargin']},
    percent_keys=['dealerMargin','msrpMargin'], top_levels=SW_TOP,
    expect={1:'Model Number',2:'Description',3:'HSN',4:'P_ListPrice',5:'Dealer Price',
            6:'Sub dealer price',7:'Distributor Price Tax In',8:'MSRP (P_Price',
            9:'MSRP - 35%',10:'MSRP - 30%',11:'MSRP - 20%',12:'MSRP - 15%',
            13:'Dealer price msrp',14:'MSRP margin'})
praveen.update({'id':'praveen','label':'Price List','internalOnly':False})

distdealer = build_list(swwb['Stonewater_Dist_Dealer_price'], header_row=2, model=0, desc=1, hsn=2,
    fields={'distCost':(3,'inr'),'listPrice':(4,'inr'),'dealer':(5,'inr'),'subdealer':(6,'inr'),
            'distInclTax':(7,'inr'),'msrp':(8,'inr'),'msrp30':(9,'inr'),'msrp20':(10,'inr'),
            'dealerDistMargin':(11,'pct'),'subdealerDistMargin':(12,'pct')},
    labels={'distCost':'Distributor cost','listPrice':'List Price +18%','dealer':'Dealer',
            'subdealer':'Sub-dealer','distInclTax':'Distributor (tax incl.)','msrp':'MSRP',
            'msrp30':'MSRP −30%','msrp20':'MSRP −20%','dealerDistMargin':'Dealer-dist margin',
            'subdealerDistMargin':'Sub-dealer-dist margin'},
    roles={'customer':['msrp'],'dealer':['dealer','msrp'],'subdealer':['subdealer','msrp'],
           'internal':['msrp','msrp30','msrp20','dealer','subdealer','listPrice','distCost','distInclTax','dealerDistMargin','subdealerDistMargin']},
    percent_keys=['dealerDistMargin','subdealerDistMargin'], top_levels=SW_TOP,
    expect={0:'Model Number',1:'Description',2:'HSN',3:'Distributor Price',
            4:'List_Price + tax',5:'Dealer Price',6:'Sub dealer price',
            7:'Distributor Price Tax In',8:'MSRP inclusive',9:'MSRP - 30%',
            10:'MSRP - 20%',11:'Dealer-dist margin',12:'Subdealer-dist margin'})
distdealer.update({'id':'distdealer','label':'Dist / Dealer','internalOnly':True})

# ---------------- Kasper ----------------
kawb = load_workbook(KA, data_only=True)
kasper = build_list(kawb['Products'], header_row=0, model=1, desc=2, hsn=None,
    fields={'distRga':(3,'inr'),'distInclTax':(4,'inr'),'dealer':(5,'inr'),
            'listPlusTax':(6,'inr'),'mrp':(7,'inr'),
            'dealerMargin':(8,'pct'),'distMargin':(9,'pct')},
    labels={'distRga':'Dist RGA cost','distInclTax':'Distributor (tax incl.)',
            'dealer':'Dealer','listPlusTax':'List price +Tax',
            'mrp':'MRP','dealerMargin':'Dealer margin','distMargin':'Dist margin'},
    roles={'customer':['mrp'],'dealer':['dealer','mrp'],'subdealer':['mrp'],
           'internal':['mrp','dealer','listPlusTax','distRga','distInclTax','dealerMargin','distMargin']},
    percent_keys=['dealerMargin','distMargin'], top_levels=None,
    expect={1:'Model No',2:'Description',3:'Dist RGA Price',
            4:'Dist RGA Price (inc taxes)',5:'Dealer Price',6:'List price plus Tax',
            7:'MRP',8:'Dealer Margin',9:'Dist Margin'})
kasper.update({'id':'products','label':'Price List','internalOnly':False})

out = {'brands':[
    {'id':'stonewater','name':'Stonewater','effectiveDate':'01 Apr 2026','lists':[praveen,distdealer]},
    {'id':'kasper','name':'Kasper','effectiveDate':'2026','lists':[kasper]},
]}
os.makedirs(SITE_DIR, exist_ok=True)
with open(os.path.join(SITE_DIR, 'prices.json'), 'w') as f:
    json.dump(out, f, ensure_ascii=False, separators=(',',':'))

for br in out['brands']:
    print('BRAND', br['name'], '| eff', br['effectiveDate'])
    for L in br['lists']:
        n = sum(len(s['products']) for c in L['categories'] for s in c['subcategories'])
        print(f"   list '{L['label']}' (id={L['id']}, internalOnly={L['internalOnly']}): "
              f"{len(L['categories'])} cats, {sum(len(c['subcategories']) for c in L['categories'])} subcats, {n} products")
